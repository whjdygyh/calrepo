"""Generate CalRepo's complete General Roman Calendar from Apache-2.0 upstream data."""

from __future__ import annotations

import json
from datetime import date, datetime, timedelta
from pathlib import Path
from urllib.request import Request, urlopen

from openpyxl import load_workbook

YEAR_URL = "https://litcal.johnromanodorazio.com/api/v5/calendar?year={year}&year_type=CIVIL&locale=en"
MISSAL_BASE = "https://raw.githubusercontent.com/Liturgical-Calendar/LiturgicalCalendarAPI/stable/jsondata/sourcedata/missals"
FIXED_RULE_URLS = tuple(f"{MISSAL_BASE}/propriumdesanctis_{year}/propriumdesanctis_{year}.json" for year in (1970, 2002, 2008))
TRANSLATION_URLS = tuple(f"{MISSAL_BASE}/propriumdesanctis_{year}/i18n/en.json" for year in (1970, 2002, 2008))


def fetch_json(url: str) -> object:
    request = Request(url, headers={"User-Agent": "CalRepo/1.0 (+https://calrepo.com)"})
    with urlopen(request, timeout=30) as response:
        return json.load(response)


def escape_ical(value: str) -> str:
    return value.replace("\\", "\\\\").replace(";", "\\;").replace(",", "\\,").replace("\n", "\\n")


def complete_year(year: int, fixed_rules: dict[str, dict[str, object]], translations: dict[str, str]) -> list[tuple[date, str, str, int]]:
    payload = fetch_json(YEAR_URL.format(year=year))
    annual_events = payload.get("litcal", []) if isinstance(payload, dict) else payload
    events: dict[tuple[date, str], tuple[str, int]] = {}
    active_keys: set[str] = set()
    for item in annual_events:
        if not isinstance(item, dict):
            continue
        event_date = date.fromisoformat(str(item.get("date", ""))[:10])
        name = str(item.get("name", "")).strip()
        event_key = str(item.get("event_key", ""))
        if event_date.year != year or not name:
            continue
        if event_key:
            active_keys.add(event_key)
        events[(event_date, name)] = (event_key, int(item.get("grade", 0) or 0))
    for event_key, rule in fixed_rules.items():
        if event_key in active_keys:
            continue
        try:
            event_date = date(year, int(rule["month"]), int(rule["day"]))
        except (KeyError, TypeError, ValueError):
            continue
        name = translations.get(event_key, "").strip()
        if name:
            events[(event_date, name)] = (event_key, 0)
    return [(event_date, name, event_key, grade) for (event_date, name), (event_key, grade) in sorted(events.items())]


def load_lay_rules() -> tuple[int, set[str], tuple[str, ...]]:
    workbook = load_workbook(Path("data/seed/calendar-rules.xlsx"), read_only=True, data_only=True)
    sheet = workbook["天主教罗马礼规则"]
    minimum_grade, core_keys, excluded_suffixes = 99, set(), []
    for rule_id, rule_type, value, lay_core, *_ in sheet.iter_rows(min_row=2, values_only=True):
        if lay_core != "是":
            continue
        if rule_type == "minimum_grade":
            minimum_grade = int(value)
        elif rule_type == "event_key":
            core_keys.add(str(value))
    for _, rule_type, value, *_ in sheet.iter_rows(min_row=2, values_only=True):
        if rule_type == "exclude_key_suffix":
            excluded_suffixes.append(str(value))
    return minimum_grade, core_keys, tuple(excluded_suffixes)


def main() -> None:
    fixed_rules: dict[str, dict[str, object]] = {}
    for url in FIXED_RULE_URLS:
        payload = fetch_json(url)
        if isinstance(payload, list):
            fixed_rules.update({str(rule["event_key"]): rule for rule in payload if isinstance(rule, dict) and rule.get("event_key")})
    translations: dict[str, str] = {}
    for url in TRANSLATION_URLS:
        payload = fetch_json(url)
        if isinstance(payload, dict):
            translations.update({str(key): str(value) for key, value in payload.items()})
    events = [event for year in range(2025, 2036) for event in complete_year(year, fixed_rules, translations)]
    minimum_grade, core_keys, excluded_suffixes = load_lay_rules()
    lay_events = [event for event in events if not event[2].endswith(excluded_suffixes) and (event[3] >= minimum_grade or event[2] in core_keys)]
    lines = ["BEGIN:VCALENDAR", "VERSION:2.0", "PRODID:-//CalRepo//General Roman Calendar//EN", "CALSCALE:GREGORIAN", "METHOD:PUBLISH", "X-WR-CALNAME:General Roman Calendar", "X-WR-TIMEZONE:UTC", "X-CALREPO-SOURCE-ID:SRC-CATHOLIC", "X-CALREPO-VERSION:2026.07.25"]
    for event_date, name, event_key, grade in events:
        key = "".join(character.lower() if character.isalnum() else "-" for character in event_key or name).strip("-")
        core = (event_date, name, event_key, grade) in lay_events
        categories = "Catholic,General Roman Calendar,Core Lay Observance" if core else "Catholic,General Roman Calendar"
        lines.extend(["BEGIN:VEVENT", f"UID:calrepo.catholic.general-roman.{key}.{event_date:%Y%m%d}@calrepo.com", "DTSTAMP:20260726T000000Z", f"DTSTART;VALUE=DATE:{event_date:%Y%m%d}", f"DTEND;VALUE=DATE:{event_date + timedelta(days=1):%Y%m%d}", f"SUMMARY:{escape_ical(name)}", "DESCRIPTION:General Roman Calendar observance. Includes annual liturgies and universal fixed observances; excludes national, diocesan, and religious-order calendars.", f"CATEGORIES:{categories}", f"X-CALREPO-LAY-PRIORITY:{'core' if core else 'ordinary'}", "END:VEVENT"])
    lines.extend(["END:VCALENDAR", ""])
    output = Path("ical/v1/religion/catholic-general-roman/en.ics")
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_bytes("\r\n".join(lines).encode("utf-8"))
    lay_lines = ["BEGIN:VCALENDAR", "VERSION:2.0", "PRODID:-//CalRepo//Catholic Lay Core Calendar//EN", "CALSCALE:GREGORIAN", "METHOD:PUBLISH", "X-WR-CALNAME:Catholic Lay Core Calendar", "X-CALREPO-SOURCE-ID:SRC-CATHOLIC-LAITY-CORE"]
    for event_date, name, event_key, _ in lay_events:
        key = "".join(character.lower() if character.isalnum() else "-" for character in event_key or name).strip("-")
        lay_lines.extend(["BEGIN:VEVENT", f"UID:calrepo.catholic.lay-core.{key}.{event_date:%Y%m%d}@calrepo.com", "DTSTAMP:20260726T000000Z", f"DTSTART;VALUE=DATE:{event_date:%Y%m%d}", f"DTEND;VALUE=DATE:{event_date + timedelta(days=1):%Y%m%d}", f"SUMMARY:{escape_ical(name)}", "CATEGORIES:Catholic,Lay Core Observance", "X-CALREPO-LAY-PRIORITY:core", "END:VEVENT"])
    lay_lines.extend(["END:VCALENDAR", ""])
    lay_output = Path("ical/v1/religion/catholic-lay-core/en.ics")
    lay_output.parent.mkdir(parents=True, exist_ok=True)
    lay_output.write_bytes("\r\n".join(lay_lines).encode("utf-8"))
    print(f"generated {len(events)} complete and {len(lay_events)} lay-core events")


if __name__ == "__main__":
    main()
